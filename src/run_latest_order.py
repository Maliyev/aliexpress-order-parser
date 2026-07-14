"""Process the newest HTML file saved by the Chrome extension."""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from parse_order import ensure_empty, next_order_row, parse_order, write_order_block


PROJECT_ROOT = Path(__file__).resolve().parents[1]


def load_config(config_path: Path) -> dict:
    if not config_path.exists():
        raise FileNotFoundError(
            f"Config file not found: {config_path}\n"
            "Copy config.example.json to config.json and set the local paths."
        )

    config = json.loads(config_path.read_text(encoding="utf-8"))
    required = ["downloads_folder", "excel_template", "sheet", "output_folder", "backup_folder"]
    missing = [key for key in required if not config.get(key)]
    if missing:
        raise ValueError("Missing config values: " + ", ".join(missing))
    return config


def find_latest_download(downloads_folder: Path) -> Path:
    files = list(downloads_folder.glob("AliExpress-order-*.html"))
    if not files:
        raise FileNotFoundError(
            "No AliExpress-order-*.html file was found in Downloads. "
            "Open an order, use the Chrome extension, then try again."
        )
    return max(files, key=lambda file: file.stat().st_mtime)


def order_digits(value) -> str:
    return "".join(re.findall(r"\d", str(value or "")))


def find_duplicates(sheet, order) -> list[str]:
    duplicates = []
    for row_number, row in enumerate(
        sheet.iter_rows(min_row=2, max_col=6, values_only=True), start=2
    ):
        tracking = str(row[3] or "").strip()
        saved_order_number = order_digits(row[5])

        if tracking == order.tracking_number:
            duplicates.append(f"tracking number in row {row_number}")
        if saved_order_number == order.order_number:
            duplicates.append(f"order ID in row {row_number}")
    return duplicates


def show_preview(order, html_path: Path):
    print("\nLatest file:", html_path.name)
    print("Order ID:", order.order_number)
    print("Tracking:", order.tracking_number)
    print("Seller:", order.seller)
    print("Products:", len(order.products))
    for number, product in enumerate(order.products, start=1):
        variation = f" | {product.variation}" if product.variation else ""
        print(f"  {number}. {product.quantity} x ${product.price:.2f}{variation} | {product.title}")
    print(f"Total: ${order.total_price:.2f}")


def ask_value(label: str, default: str) -> str:
    answer = input(f"{label} [{default}]: ").strip()
    return answer or default


def make_backup(workbook, source_path: Path, backup_folder: Path) -> Path:
    backup_folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_path = backup_folder / f"{source_path.stem}-{timestamp}.xlsx"
    workbook.save(backup_path)
    return backup_path


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", type=Path, default=PROJECT_ROOT / "config.json")
    args = parser.parse_args()

    config = load_config(args.config)
    html_path = find_latest_download(Path(config["downloads_folder"]))
    order = parse_order(html_path)
    show_preview(order, html_path)

    defaults = config.get("manual_defaults", {})
    print("\nThese fields are not in the saved AliExpress page. Leave blank if unknown.")
    arrival_date = ask_value("Column B, arrival date", defaults.get("arrival_date", ""))
    expected_arrival = ask_value("Column C, expected arrival", defaults.get("expected_arrival", ""))
    forwarder_code = ask_value("Column E, forwarder or SF code", defaults.get("forwarder_code", ""))

    source_path = Path(config["excel_template"])
    workbook = load_workbook(source_path)
    if config["sheet"] not in workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {config['sheet']}")
    sheet = workbook[config["sheet"]]

    duplicates = find_duplicates(sheet, order)
    if duplicates:
        print("\nStopped. This order may already be in the workbook:")
        for item in duplicates:
            print("-", item)
        return

    answer = input("\nCreate a backup and a new Excel copy? [y/N]: ").strip().lower()
    if answer not in {"y", "yes"}:
        print("Nothing was written.")
        return

    backup_path = make_backup(workbook, source_path, Path(config["backup_folder"]))
    start_row = next_order_row(sheet)
    row_count = 1 + sum(
        3 + bool(product.variation) + (product.quantity > 1)
        for product in order.products
    ) + 5
    ensure_empty(sheet, start_row, row_count)
    end_row = write_order_block(
        sheet,
        sheet,
        order,
        start_row,
        arrival_date,
        expected_arrival,
        forwarder_code,
    )

    output_folder = Path(config["output_folder"])
    output_folder.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    output_path = output_folder / f"order-{order.order_number}-{timestamp}.xlsx"
    workbook.save(output_path)

    print("\nBackup:", backup_path)
    print("New Excel copy:", output_path)
    print(f"Rows written: {start_row}-{end_row}")


if __name__ == "__main__":
    main()
