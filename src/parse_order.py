"""Read one saved AliExpress order page and add its block to an Excel copy."""

from __future__ import annotations

import argparse
import json
import re
from copy import copy
from dataclasses import asdict, dataclass
from datetime import date
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import Workbook, load_workbook


@dataclass
class Product:
    title: str
    variation: str | None
    quantity: int
    price: float
    product_url: str


@dataclass
class Order:
    order_date: str
    order_date_serial: int
    order_number: str
    tracking_number: str
    seller: str
    seller_url: str
    products: list[Product]
    total_price: float
    item_count_label: str
    item_subtotal: float
    delivery_label: str
    delivery_price: float


class Node:
    def __init__(self, tag: str = "", attrs: list[tuple[str, str | None]] | None = None):
        self.tag = tag
        self.attrs = dict(attrs or [])
        self.children: list[Node | str] = []


class PageTree(HTMLParser):
    def __init__(self):
        super().__init__(convert_charrefs=True)
        self.root = Node("root")
        self.stack = [self.root]

    def handle_starttag(self, tag, attrs):
        node = Node(tag, attrs)
        self.stack[-1].children.append(node)
        self.stack.append(node)

    def handle_startendtag(self, tag, attrs):
        self.stack[-1].children.append(Node(tag, attrs))

    def handle_endtag(self, tag):
        for index in range(len(self.stack) - 1, 0, -1):
            if self.stack[index].tag == tag:
                self.stack = self.stack[:index]
                return

    def handle_data(self, data):
        self.stack[-1].children.append(data)


def walk(node: Node):
    yield node
    for child in node.children:
        if isinstance(child, Node):
            yield from walk(child)


def raw_text(node: Node) -> str:
    parts = []
    for child in node.children:
        if isinstance(child, str):
            parts.append(child)
        else:
            parts.append(raw_text(child))
    return "".join(parts)


def text(node: Node) -> str:
    return " ".join(raw_text(node).split())


def class_has(node: Node, part: str) -> bool:
    return part in node.attrs.get("class", "")


def first_descendant(node: Node, predicate, description: str) -> Node:
    for item in walk(node):
        if predicate(item):
            return item
    raise ValueError(f"Could not find {description} in the saved page.")


def money(value: str) -> float:
    match = re.search(r"\$\s*([0-9]+(?:[.,][0-9]+)?)", value)
    if not match:
        raise ValueError(f"Could not read a dollar amount from: {value!r}")
    return float(match.group(1).replace(",", "."))


MONTHS = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
    "января": 1,
    "февраля": 2,
    "марта": 3,
    "апреля": 4,
    "мая": 5,
    "июня": 6,
    "июля": 7,
    "августа": 8,
    "сентября": 9,
    "октября": 10,
    "ноября": 11,
    "декабря": 12,
    "yanvar": 1,
    "fevral": 2,
    "mart": 3,
    "aprel": 4,
    "may": 5,
    "iyun": 6,
    "iyul": 7,
    "avqust": 8,
    "sentyabr": 9,
    "oktyabr": 10,
    "noyabr": 11,
    "dekabr": 12,
}

ORDER_DATE_PATTERN = re.compile(
    r"(?:order\s+date|date|дата\s+заказа|дата|sifariş\s+tarixi)\s*:?\s*"
    r"(\d{1,2})\s+([A-Za-zА-Яа-яЁё]+)\s+(\d{4})",
    re.IGNORECASE,
)


def parse_order_date(root: Node) -> date:
    for node in walk(root):
        candidate = text(node)
        if len(candidate) > 120:
            continue

        match = ORDER_DATE_PATTERN.search(candidate)
        if not match:
            continue

        day, month_name, year = match.groups()
        month = MONTHS.get(month_name.lower())
        if month:
            return date(int(year), month, int(day))

    raise ValueError(
        "Could not find an order date. The parser supports English, Russian, and Azerbaijani order-date labels."
    )


def summary_from_page_data(html: str) -> dict:
    marker = '"summary":{"total":'
    start = html.find(marker)
    if start == -1:
        raise ValueError("Could not find the order summary data in the saved page.")

    summary, _ = json.JSONDecoder().raw_decode(html[start + len('"summary":'):])
    return summary["total"]


def parse_order(html_path: Path) -> Order:
    html = html_path.read_text(encoding="utf-8")
    parser = PageTree()
    parser.feed(html)

    desktop_order_number = first_descendant(
        parser.root,
        lambda node: node.attrs.get("data-testid") == "orderNumber" and bool(re.search(r"\d", text(node))),
        "desktop order number",
    )
    order_number = "".join(re.findall(r"\d", text(desktop_order_number)))

    order_day = parse_order_date(parser.root)
    order_date_serial = (order_day - date(1899, 12, 30)).days

    seller_block = first_descendant(
        parser.root,
        lambda node: node.attrs.get("data-testid") == "sellerInfoV2Block",
        "seller information",
    )
    seller = re.sub(r"^Seller:\s*", "", text(seller_block))
    seller_link = first_descendant(seller_block, lambda node: node.tag == "a", "seller link")

    tracking_text = first_descendant(
        parser.root,
        lambda node: class_has(node, "InfoDeliveryButton__trackNumber__"),
        "tracking number",
    )
    tracking_number = "".join(re.findall(r"[A-Za-z0-9]", raw_text(tracking_text)))

    products = []
    for block in walk(parser.root):
        if block.attrs.get("data-testid") != "productBlock":
            continue

        title = text(first_descendant(block, lambda node: class_has(node, "Product__title__"), "product title"))
        variation_nodes = [node for node in walk(block) if class_has(node, "Product__description__")]
        variation = text(variation_nodes[0]) if variation_nodes else None

        quantity_text = text(
            first_descendant(block, lambda node: class_has(node, "Product__quantityText__"), "product quantity")
        )
        quantity_match = re.search(r"(\d+)\s*pcs?\.", quantity_text)
        if not quantity_match:
            raise ValueError(f"Could not read product quantity from: {quantity_text!r}")

        price_box = first_descendant(
            block,
            lambda node: class_has(node, "Product__priceDesktop__"),
            "desktop product price",
        )
        price_nodes = [child for child in price_box.children if isinstance(child, Node)]
        price_text = next(
            (text(node) for node in price_nodes if re.fullmatch(r"\$\s*[0-9]+(?:[.,][0-9]+)?", text(node))),
            None,
        )
        if not price_text:
            raise ValueError(f"Could not read product price for: {title!r}")

        product_link = first_descendant(
            block,
            lambda node: node.attrs.get("data-testid") == "productText" and node.tag == "a",
            "product link",
        )
        products.append(
            Product(
                title=title,
                variation=variation,
                quantity=int(quantity_match.group(1)),
                price=money(price_text),
                product_url=product_link.attrs["href"],
            )
        )

    summary = summary_from_page_data(html)
    rows = summary.get("list", [])
    if len(rows) < 2:
        raise ValueError("Could not find the item subtotal and delivery rows.")

    item_row, delivery_row = rows[:2]
    item_count_label = item_row["title"]
    delivery_label = delivery_row["title"]

    return Order(
        order_date=order_day.isoformat(),
        order_date_serial=order_date_serial,
        order_number=order_number,
        tracking_number=tracking_number,
        seller=seller,
        seller_url=seller_link.attrs["href"],
        products=products,
        total_price=money(summary["value"]),
        item_count_label=item_count_label,
        item_subtotal=money(item_row["value"]),
        delivery_label=delivery_label,
        delivery_price=money(delivery_row["value"]),
    )


def copy_style(source, target):
    target._style = copy(source._style)
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)
    target.number_format = source.number_format
    target.hyperlink = None


def copy_row_height(source_sheet, target_sheet, source_row: int, target_row: int):
    target_sheet.row_dimensions[target_row].height = source_sheet.row_dimensions[source_row].height


def next_order_row(sheet) -> int:
    last_nonempty = 1
    for row in sheet.iter_rows():
        if any(cell.value is not None for cell in row):
            last_nonempty = row[0].row
    return last_nonempty + 3


def block_length(order: Order) -> int:
    product_rows = sum(3 + (1 if product.variation else 0) + (1 if product.quantity > 1 else 0) for product in order.products)
    return 1 + product_rows + 5


def ensure_empty(sheet, start_row: int, rows: int):
    occupied = []
    for row in sheet.iter_rows(min_row=start_row, max_row=start_row + rows - 1, min_col=1, max_col=7):
        for cell in row:
            if cell.value is not None:
                occupied.append(cell.coordinate)
    if occupied:
        raise ValueError("The target block is not empty: " + ", ".join(occupied[:10]))


def write_order_block(sheet, style_sheet, order: Order, start_row: int, arrival_date: str | None, expected_arrival: str | None, forwarder_code: str | None):
    reference_row = 1055
    for column in range(1, 8):
        copy_style(style_sheet.cell(reference_row, column), sheet.cell(start_row, column))
    copy_row_height(style_sheet, sheet, reference_row, start_row)

    values = [
        order.order_date_serial,
        arrival_date or None,
        expected_arrival or None,
        order.tracking_number,
        forwarder_code or None,
        order.order_number,
        order.seller,
    ]
    for column, value in enumerate(values, start=1):
        sheet.cell(start_row, column).value = value

    if not forwarder_code:
        copy_style(style_sheet["E1053"], sheet.cell(start_row, 5))

    row = start_row + 1
    for product in order.products:
        if product.quantity > 1:
            copy_style(style_sheet["F1056"], sheet.cell(row, 6))
            copy_row_height(style_sheet, sheet, 1056, row)
            sheet.cell(row, 6).value = product.quantity
            row += 1

        copy_style(style_sheet["F1057"], sheet.cell(row, 6))
        copy_row_height(style_sheet, sheet, 1057, row)
        sheet.cell(row, 6).value = product.title
        row += 1

        if product.variation:
            copy_style(style_sheet["F1058"], sheet.cell(row, 6))
            copy_row_height(style_sheet, sheet, 1058, row)
            sheet.cell(row, 6).value = product.variation
            row += 1

        copy_style(style_sheet["F1059"], sheet.cell(row, 6))
        copy_row_height(style_sheet, sheet, 1059, row)
        sheet.cell(row, 6).value = product.price
        row += 1

        copy_style(style_sheet["F1060"], sheet.cell(row, 6))
        copy_row_height(style_sheet, sheet, 1060, row)
        sheet.cell(row, 6).value = f"{product.quantity} {'pc.' if product.quantity == 1 else 'pcs.'}"
        row += 1

    summary_values = [
        ("F1102", f"Total price:$ {order.total_price:.2f}"),
        ("F1103", order.item_count_label),
        ("F1104", order.item_subtotal),
        ("F1105", order.delivery_label),
        ("F1106", order.delivery_price),
    ]
    for source, value in summary_values:
        copy_style(style_sheet[source], sheet.cell(row, 6))
        copy_row_height(style_sheet, sheet, style_sheet[source].row, row)
        sheet.cell(row, 6).value = value
        row += 1

    return row - 1


def make_empty_test_workbook(style_sheet):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = style_sheet.title
    sheet.sheet_view.showGridLines = style_sheet.sheet_view.showGridLines

    for column in "ABCDEFG":
        source_dimension = style_sheet.column_dimensions[column]
        target_dimension = sheet.column_dimensions[column]
        target_dimension.width = source_dimension.width
        target_dimension.hidden = source_dimension.hidden

    for column in range(1, 8):
        copy_style(style_sheet.cell(1, column), sheet.cell(1, column))
        sheet.cell(1, column).value = style_sheet.cell(1, column).value
    copy_row_height(style_sheet, sheet, 1, 1)
    return workbook, sheet


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("html", type=Path, help="Saved AliExpress order HTML file")
    parser.add_argument("--workbook", type=Path, help="Original Excel workbook to copy")
    parser.add_argument("--output", type=Path, help="Path for the new Excel copy")
    parser.add_argument("--sheet", default="Лист2", help="Worksheet name, default: Лист2")
    parser.add_argument("--start-row", type=int, help="First row for the new block; default: next free block")
    parser.add_argument("--new-workbook", action="store_true", help="Create an otherwise empty workbook with one test order")
    parser.add_argument("--arrival-date", help="Manual value for column B")
    parser.add_argument("--expected-arrival", help="Manual value for column C")
    parser.add_argument("--forwarder-code", help="Manual value for column E")
    parser.add_argument("--json", type=Path, help="Optional path for extracted data as JSON")
    args = parser.parse_args()

    order = parse_order(args.html)
    payload = asdict(order)
    if args.json:
        args.json.parent.mkdir(parents=True, exist_ok=True)
        args.json.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"Order: {order.order_number}")
    print(f"Tracking: {order.tracking_number}")
    print(f"Products: {len(order.products)}")
    print("UNRESOLVED from HTML: columns B, C and E")

    if not args.workbook and not args.output:
        return
    if not args.workbook or not args.output:
        raise ValueError("Use both --workbook and --output, or neither.")
    if args.workbook.resolve() == args.output.resolve():
        raise ValueError("Output must be a new file, not the original workbook.")
    if args.output.exists():
        raise FileExistsError(f"Output already exists: {args.output}")

    template_workbook = load_workbook(args.workbook)
    if args.sheet not in template_workbook.sheetnames:
        raise ValueError(f"Worksheet not found: {args.sheet}")
    style_sheet = template_workbook[args.sheet]

    if args.new_workbook:
        workbook, sheet = make_empty_test_workbook(style_sheet)
        start_row = args.start_row or 3
    else:
        workbook = template_workbook
        sheet = style_sheet
        start_row = args.start_row or next_order_row(sheet)
        ensure_empty(sheet, start_row, block_length(order))

    end_row = write_order_block(
        sheet,
        style_sheet,
        order,
        start_row,
        args.arrival_date,
        args.expected_arrival,
        args.forwarder_code,
    )

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Created: {args.output}")
    print(f"Excel rows: {start_row}-{end_row}")


if __name__ == "__main__":
    main()
