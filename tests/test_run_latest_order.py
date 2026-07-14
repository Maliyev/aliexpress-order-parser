import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "src"))

from run_latest_order import make_backup, make_room_at_top


class BackupTests(unittest.TestCase):
    def test_backup_is_saved_from_open_workbook(self):
        workbook = Workbook()
        workbook.active["A1"] = "keep this value"

        with tempfile.TemporaryDirectory() as folder:
            folder_path = Path(folder)
            backup_path = make_backup(
                workbook,
                folder_path / "orders.xlsx",
                folder_path / "backups",
            )

            self.assertTrue(backup_path.exists())
            self.assertEqual(load_workbook(backup_path).active["A1"].value, "keep this value")

    def test_new_order_rows_are_inserted_at_row_51(self):
        sheet = Workbook().active
        sheet["A51"] = "existing order"
        sheet["F1055"] = "style source"

        start_row, reference_row = make_room_at_top(sheet, 8)

        self.assertEqual(start_row, 51)
        self.assertEqual(reference_row, 1063)
        self.assertEqual(sheet["A59"].value, "existing order")
        self.assertEqual(sheet["F1063"].value, "style source")
